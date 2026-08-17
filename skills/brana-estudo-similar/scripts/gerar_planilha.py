#!/usr/bin/env python3
"""
Gera a planilha comparativa do estudo de similares.

Uso:
    python gerar_planilha.py dados.json [-o pasta_de_saida]
    python gerar_planilha.py --exemplo          # imprime um dados.json de exemplo

Saida: "Estudo de Similar - <Projetista>.xlsx", com quatro abas:

  Metrico (SI)     todos os valores em m, kg, L, kW, nos - SO FORMULA, nada
                   digitado. E a aba que se le.
  Original (fonte) o valor exatamente como a fonte publicou, com a unidade de
                   origem declarada por coluna. E a aba que se audita.
  Fontes           link, data de publicacao e data de acesso de cada pagina.
  Fatores          os fatores de conversao usados, com observacao.

A separacao entre "original" e "metrico" e o que torna o estudo defensavel: o
cliente abre a aba original, compara com o site e confirma que ninguem mexeu no
numero. Por isso a aba metrica nunca recebe valor digitado.

Requer openpyxl.
"""

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl nao encontrado. Rode: pip install openpyxl --break-system-packages")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from parametros import (  # noqa: E402
    FATORES, FATOR_LINHA, PARAMETROS, RAZOES, valida_similar,
)

AZUL = "1F3864"
CINZA = "D9D9D9"
CINZA_CLARO = "F2F2F2"
AMARELO = "FFF2CC"

FINA = Side(style="thin", color="BFBFBF")
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)

FMT = {
    "FT_M": "0.00", "FT2_M2": "0.00", "LB_KG": "#,##0", "GAL_L": "#,##0",
    "HP_KW": "#,##0", "MPH_KN": "0.0",
}


def slug(txt):
    txt = unicodedata.normalize("NFKD", str(txt)).encode("ascii", "ignore").decode()
    txt = re.sub(r"[^\w\s-]", "", txt).strip()
    return re.sub(r"\s+", " ", txt) or "Projetista"


def titulo(ws, cel, texto, tam=12):
    ws[cel] = texto
    ws[cel].font = Font(bold=True, size=tam, color=AZUL)


def cabecalho(cel):
    cel.font = Font(bold=True, color="FFFFFF")
    cel.fill = PatternFill("solid", fgColor=AZUL)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = BORDA


# --------------------------------------------------------------------------
# aba Original (fonte)
# --------------------------------------------------------------------------
LIN_CAB = 4          # linha do cabecalho
LIN_SIST = 5         # linha que declara o sistema de unidades da fonte
LIN_INI = 7          # primeiro parametro
COL_INI = 6          # coluna F = primeira embarcacao


def aba_original(ws, similares):
    titulo(ws, "B1", "Estudo de similares - dados como publicados pela fonte")
    ws["B2"] = ("Cada valor abaixo foi copiado literalmente da pagina indicada na aba "
                "Fontes, sem arredondar. A unidade depende do sistema declarado na "
                "linha 5 de cada coluna. Celula vazia = a fonte nao informa.")
    ws["B2"].font = Font(italic=True, size=9)

    for col, rot in ((2, "Parametro (PT)"), (3, "Parameter (EN)"),
                     (4, "Unid. se imperial"), (5, "Unid. se metrico")):
        c = ws.cell(row=LIN_CAB, column=col, value=rot)
        cabecalho(c)
    for i, sim in enumerate(similares):
        c = ws.cell(row=LIN_CAB, column=COL_INI + i, value=nome_completo(sim))
        cabecalho(c)

    ws.cell(row=LIN_SIST, column=2, value="Sistema de unidades da fonte")
    ws.cell(row=LIN_SIST, column=2).font = Font(bold=True, italic=True)
    for i, sim in enumerate(similares):
        c = ws.cell(row=LIN_SIST, column=COL_INI + i, value=sim.get("sistema", ""))
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=AMARELO)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDA

    linhas = {}
    for j, p in enumerate(PARAMETROS):
        r = LIN_INI + j
        linhas[p["key"]] = r
        ws.cell(row=r, column=2, value=p["pt"])
        ws.cell(row=r, column=3, value=p["en"])
        ws.cell(row=r, column=4, value=p["unid_imp"])
        ws.cell(row=r, column=5, value=p["unid_si"])
        for col in (2, 3, 4, 5):
            ws.cell(row=r, column=col).border = BORDA
            ws.cell(row=r, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 2))
        if j % 2:
            for col in range(2, COL_INI + len(similares)):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CINZA_CLARO)
        for i, sim in enumerate(similares):
            v = (sim.get("dados") or {}).get(p["key"])
            c = ws.cell(row=r, column=COL_INI + i)
            if v is not None:
                c.value = v
            c.border = BORDA
            c.alignment = Alignment(horizontal="center" if p["tipo"] == "num" else "left",
                                    vertical="center", wrap_text=p["tipo"] == "texto")

    # razoes publicadas pela fonte (quando houver) - ficam separadas
    r = LIN_INI + len(PARAMETROS) + 1
    ws.cell(row=r, column=2, value="RAZOES publicadas pela fonte (em branco = calculada)")
    ws.cell(row=r, column=2).font = Font(bold=True, color=AZUL)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=CINZA)
    for j, razao in enumerate(RAZOES):
        rr = r + 1 + j
        linhas[razao["key"]] = rr
        ws.cell(row=rr, column=2, value=razao["pt"])
        ws.cell(row=rr, column=3, value=razao["en"])
        ws.cell(row=rr, column=4, value="-")
        ws.cell(row=rr, column=5, value=razao["unid_si"])
        for col in (2, 3, 4, 5):
            ws.cell(row=rr, column=col).border = BORDA
        for i, sim in enumerate(similares):
            v = (sim.get("dados") or {}).get(razao["key"])
            c = ws.cell(row=rr, column=COL_INI + i)
            if v is not None:
                c.value = v
            c.border = BORDA
            c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    for i in range(len(similares)):
        ws.column_dimensions[get_column_letter(COL_INI + i)].width = 22
    ws.freeze_panes = ws.cell(row=LIN_INI, column=COL_INI)
    return linhas


# --------------------------------------------------------------------------
# aba Metrico (SI) - so formula
# --------------------------------------------------------------------------
def aba_metrica(ws, similares, orig_linhas, nome_orig):
    titulo(ws, "B1", "Comparativo de similares - dimensoes no sistema metrico (SI)")
    ws["B2"] = (f"Valores calculados por formula a partir da aba '{nome_orig}' "
                f"(dados originais preservados). Fatores na aba 'Fatores'. "
                f"Nenhum valor desta aba foi digitado.")
    ws["B2"].font = Font(italic=True, size=9)

    for col, rot in ((2, "Parametro (PT)"), (3, "Parameter (EN)"), (4, "Unidade")):
        cabecalho(ws.cell(row=LIN_CAB, column=col, value=rot))
    COL_M = 5
    for i, sim in enumerate(similares):
        cabecalho(ws.cell(row=LIN_CAB, column=COL_M + i, value=nome_completo(sim)))

    o = f"'{nome_orig}'!"
    linhas = {}
    for j, p in enumerate(PARAMETROS):
        r = LIN_INI + j
        linhas[p["key"]] = r
        ws.cell(row=r, column=2, value=p["pt"])
        ws.cell(row=r, column=3, value=p["en"])
        ws.cell(row=r, column=4, value=p["unid_si"])
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDA
            ws.cell(row=r, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 2))
        if j % 2:
            for col in range(2, COL_M + len(similares)):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CINZA_CLARO)
        for i, sim in enumerate(similares):
            src = f"{o}{get_column_letter(COL_INI + i)}{orig_linhas[p['key']]}"
            imperial = sim.get("sistema") == "imperial"
            if p["fator"] and imperial:
                fator = f"Fatores!$E${FATOR_LINHA[p['fator']]}"
                f = f'=IF({src}="","",{src}*{fator})'
            else:
                f = f'=IF({src}="","",{src})'
            c = ws.cell(row=r, column=COL_M + i, value=f)
            if p["fator"]:
                c.number_format = FMT.get(p["fator"], "0.00")
            elif p["tipo"] == "num":
                c.number_format = "0"
            c.border = BORDA
            c.alignment = Alignment(horizontal="center" if p["tipo"] == "num" else "left",
                                    vertical="center", wrap_text=p["tipo"] == "texto")

    # ---- razoes: usa o valor da fonte se houver, senao calcula
    r0 = LIN_INI + len(PARAMETROS) + 1
    ws.cell(row=r0, column=2, value="RAZOES")
    ws.cell(row=r0, column=2).font = Font(bold=True, color=AZUL)
    ws.cell(row=r0, column=2).fill = PatternFill("solid", fgColor=CINZA)

    ft_m = f"Fatores!$E${FATOR_LINHA['FT_M']}"
    lt_kg = f"Fatores!$E${FATOR_LINHA['LT_KG']}"

    for j, razao in enumerate(RAZOES):
        r = r0 + 1 + j
        linhas[razao["key"]] = r
        ws.cell(row=r, column=2, value=razao["pt"])
        ws.cell(row=r, column=3, value=razao["en"])
        ws.cell(row=r, column=4, value=razao["unid_si"])
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDA
            ws.cell(row=r, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 2))
        for i, _sim in enumerate(similares):
            L = get_column_letter(COL_M + i)
            pub = f"{o}{get_column_letter(COL_INI + i)}{orig_linhas[razao['key']]}"
            loa, boa = f"{L}{linhas['loa']}", f"{L}{linhas['boa']}"
            lwl, disp = f"{L}{linhas['lwl']}", f"{L}{linhas['deslocamento']}"
            pot = f"{L}{linhas['potencia']}"
            if razao["key"] == "loa_boa":
                calc = f'IF(OR({loa}="",{boa}="",{boa}=0),"",{loa}/{boa})'
            elif razao["key"] == "dlr":
                # DLR = desloc.[long tons] / (0,01 x LWL[pes])^3, recomposto do SI
                calc = (f'IF(OR({disp}="",{lwl}="",{lwl}=0),"",'
                        f'({disp}/{lt_kg})/((0.01*({lwl}/{ft_m}))^3))')
            else:  # disp_potencia, em kg/kW
                calc = f'IF(OR({disp}="",{pot}="",{pot}=0),"",{disp}/{pot})'
            c = ws.cell(row=r, column=COL_M + i,
                        value=f'=IF({pub}<>"",{pub},{calc})')
            c.number_format = "0" if razao["key"] == "dlr" else "0.00"
            c.border = BORDA
            c.alignment = Alignment(horizontal="center")

    r = r0 + len(RAZOES) + 2
    ws.cell(row=r, column=2, value=(
        "Nota: celula vazia significa que a fonte nao informa o dado - nunca "
        "que o valor e zero. Razoes aparecem com o valor publicado pela fonte "
        "quando existe; caso contrario sao calculadas a partir das linhas acima."))
    ws.cell(row=r, column=2).font = Font(italic=True, size=9)

    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    for i in range(len(similares)):
        ws.column_dimensions[get_column_letter(COL_M + i)].width = 22
    ws.freeze_panes = ws.cell(row=LIN_INI, column=COL_M)


# --------------------------------------------------------------------------
def aba_fontes(ws, similares):
    titulo(ws, "B1", "Fontes consultadas")
    ws["B2"] = ("Somente paginas efetivamente abertas e lidas. 'sem data "
                "identificavel' significa que a pagina nao publica data.")
    ws["B2"].font = Font(italic=True, size=9)
    cols = ["Embarcacao", "Fonte", "Tipo", "Publicacao", "Acesso", "Link"]
    for i, rot in enumerate(cols):
        cabecalho(ws.cell(row=4, column=2 + i, value=rot))
    r = 5
    for sim in similares:
        for f in sim.get("fontes") or []:
            vals = [nome_completo(sim), f.get("nome", ""), f.get("tipo", ""),
                    f.get("data_publicacao", ""), f.get("data_acesso", ""),
                    f.get("url", "")]
            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=2 + i, value=v)
                c.border = BORDA
                c.alignment = Alignment(vertical="top", wrap_text=i in (0, 1))
            r += 1
        lac = sim.get("lacunas") or []
        if lac:
            c = ws.cell(row=r, column=2, value=f"{nome_completo(sim)} - lacunas: "
                                               + ", ".join(lac))
            c.font = Font(italic=True, size=9)
            r += 1
        div = sim.get("divergencias")
        if div:
            c = ws.cell(row=r, column=2, value=f"{nome_completo(sim)} - divergencia: {div}")
            c.font = Font(italic=True, size=9, color="C00000")
            r += 1
    for col, w in zip("BCDEFG", (26, 40, 14, 18, 14, 70)):
        ws.column_dimensions[col].width = w


def aba_fatores(ws):
    titulo(ws, "B2", "Fatores de conversao utilizados")
    for i, rot in enumerate(["Grandeza", "De", "Para", "Fator", "Observacao"]):
        cabecalho(ws.cell(row=4, column=2 + i, value=rot))
    for k, grandeza, de, para, valor, obs in FATORES:
        r = FATOR_LINHA[k]
        ws.cell(row=r, column=2, value=grandeza)
        ws.cell(row=r, column=3, value=de)
        ws.cell(row=r, column=4, value=para)
        if valor is None:  # derivado
            ws.cell(row=r, column=5,
                    value=f"=E{FATOR_LINHA['LB_KG']}/E{FATOR_LINHA['HP_KW']}")
        else:
            ws.cell(row=r, column=5, value=valor)
        ws.cell(row=r, column=5).number_format = "0.000000000"
        ws.cell(row=r, column=6, value=obs)
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = BORDA

    r = FATOR_LINHA[FATORES[-1][0]] + 2
    ws.cell(row=r, column=2, value="Nao convertidos (permanecem no original)")
    ws.cell(row=r, column=2).font = Font(bold=True, color=AZUL)
    notas = [
        ("Razao LOA/BOA", "Adimensional."),
        ("DLR", "Coeficiente definido em unidades imperiais (desloc. em long tons / "
                "(0,01 x LWL_pes)^3). Recalcular em SI mudaria a definicao; a formula "
                "recompoe as unidades imperiais a partir dos valores em SI."),
        ("Deadrise", "Angulo em graus - unidade comum aos dois sistemas."),
        ("Autonomia (NM)", "Milha nautica mantida (1 NM = 1,852 km)."),
        ("Motorizacao", "Texto original preservado - designacoes comerciais dos motores "
                        "(ex.: '800 HP') sao nomes de modelo."),
    ]
    for i, (a, b) in enumerate(notas):
        ws.cell(row=r + 1 + i, column=2, value=a)
        ws.cell(row=r + 1 + i, column=3, value=b)
    for col, w in zip("BCDEF", (26, 60, 12, 18, 46)):
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------------------
def nome_completo(sim):
    est = (sim.get("estaleiro") or "").strip()
    nome = (sim.get("nome") or "").strip()
    ano = (str(sim.get("ano") or "")).strip()
    txt = f"{est} {nome}".strip()
    return f"{txt} ({ano})" if ano else txt


def gerar(dados, saida):
    similares = dados.get("similares") or []
    if len(similares) < 3:
        print(f"AVISO: apenas {len(similares)} similar(es). O estudo pede no minimo 3 "
              f"- verifique se voce esgotou as fontes antes de seguir.", file=sys.stderr)
    erros = []
    for i, sim in enumerate(similares):
        erros += valida_similar(sim, i)
    if erros:
        print("Problemas nos dados:\n  - " + "\n  - ".join(erros), file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws_m = wb.active
    ws_m.title = "Metrico (SI)"
    ws_o = wb.create_sheet("Original (fonte)")
    ws_f = wb.create_sheet("Fontes")
    ws_c = wb.create_sheet("Fatores")

    orig_linhas = aba_original(ws_o, similares)
    aba_metrica(ws_m, similares, orig_linhas, ws_o.title)
    aba_fontes(ws_f, similares)
    aba_fatores(ws_c)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    saida = Path(saida)
    saida.mkdir(parents=True, exist_ok=True)
    arq = saida / f"Estudo de Similar - {slug(dados.get('projetista', ''))}.xlsx"
    wb.save(arq)
    return arq


EXEMPLO = {
    "projetista": "Ana Ribeiro",
    "data": "2026-08-17",
    "embarcacao_pretendida": "Lancha de 38 pes para pesca costeira e passeio, 8 pessoas",
    "caso_de_uso_resumo": "Resumo do caso de uso em 3 a 5 linhas.",
    "criterios_selecao": "Lanchas de 36 a 42 pes, casco em V, motorizacao de popa ou centro-rabeta.",
    "leitura_parametrica": "O que a comparacao mostra.",
    "recomendacao": "Recomendacao preliminar para o barco novo.",
    "observacoes": "Divergencias e avisos.",
    "similares": [
        {
            "nome": "400 Super Sport", "estaleiro": "Formula Boats", "ano": "2019",
            "sistema": "imperial",
            "porque_similar": "Mesma faixa de comprimento e missao de passeio costeiro.",
            "fontes": [{"nome": "Formula Boats - especificacao 400 SS",
                        "url": "https://exemplo-substituir-pelo-link-real",
                        "data_publicacao": "2019-05-10",
                        "data_acesso": "2026-08-17", "tipo": "estaleiro"}],
            "dados": {"loa": 41.5, "lwl": 33.4, "boa": 11, "calado": 3,
                      "deslocamento": 17100, "deadrise": 22, "combustivel": 250,
                      "agua": 50, "potencia": 1040, "vel_max": 47.2,
                      "vel_cruzeiro": 28.8, "autonomia": 341,
                      "motorizacao": "(2) Mercury Racing 520 Bravo Three XR"},
            "lacunas": ["pontal", "dejetos"],
        },
        {
            "nome": "One 43", "estaleiro": "Baia", "sistema": "metrico",
            "porque_similar": "Boca larga, mesma missao de passeio.",
            "fontes": [{"nome": "itBoat - Baia One 43",
                        "url": "https://exemplo-substituir-pelo-link-real",
                        "data_publicacao": "sem data identificavel",
                        "data_acesso": "2026-08-17", "tipo": "banco de dados"}],
            "dados": {"loa": 13.32, "boa": 4.6, "calado": 0.71,
                      "deslocamento": 12193, "potencia": 552,
                      "motorizacao": "(2) D6-370 Volvo Penta"},
            "lacunas": ["lwl", "deadrise", "autonomia"],
        },
        {
            "nome": "42 Sport Coupe", "estaleiro": "Regal", "sistema": "imperial",
            "porque_similar": "Comprimento equivalente, uso familiar costeiro.",
            "fontes": [{"nome": "TheBoatDB - Regal 42 Sport Coupe",
                        "url": "https://exemplo-substituir-pelo-link-real",
                        "data_publicacao": "sem data identificavel",
                        "data_acesso": "2026-08-17", "tipo": "banco de dados"}],
            "dados": {"loa": 42.3, "lwl": 30.5, "boa": 13, "calado": 3.42,
                      "deslocamento": 20500, "deadrise": 18, "potencia": 1200,
                      "vel_max": 37.2, "vel_cruzeiro": 29.7, "autonomia": 352,
                      "motorizacao": "2 x 300-hp Volvo Penta IPS400"},
            "lacunas": ["combustivel", "agua"],
        },
    ],
}


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("json", nargs="?", help="arquivo dados.json")
    ap.add_argument("-o", "--out", default=".", help="pasta de saida")
    ap.add_argument("--exemplo", action="store_true",
                    help="imprime um dados.json de exemplo e sai")
    a = ap.parse_args()
    if a.exemplo:
        print(json.dumps(EXEMPLO, indent=2, ensure_ascii=False))
        return
    if not a.json:
        ap.error("informe o dados.json (ou use --exemplo)")
    dados = json.loads(Path(a.json).read_text(encoding="utf-8"))
    print(gerar(dados, a.out))


if __name__ == "__main__":
    main()
